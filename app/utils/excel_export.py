import os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def export_excel(rows: List[Dict], out_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Name", "Email", "Phone", "Company", "Location", "Years", "URL", "CapturedAt", "Score"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("name"),
            r.get("email"),
            r.get("phone"),
            r.get("company"),
            r.get("location"),
            r.get("years_estimated"),
            r.get("url"),
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            r.get("score"),
        ])

    head_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)
    for col_idx, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")

    widths = [28, 28, 16, 24, 24, 8, 50, 20, 10]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
